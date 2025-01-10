import logging
from threading import Thread
import time
import os
from datetime import datetime
from shutil import copy2
from config import BACKUP_FOLDER, OUTPUT_FOLDER
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class ScriptManager:
    def __init__(self, socketio):
        self.socketio = socketio
        self.base_sheet_path = None
        self.new_data_sheet_path = None
        self.output_file_path = None
        self.step_functions = {
            'different_sheet_transfer': self._execute_different_sheet_transfer,
            'excel_macro': self._execute_excel_macro,
            'same_sheet_transfer': self._execute_same_sheet_transfer
        }
        self.pipeline_steps = [
            {
                'id': 'different_sheet_transfer',
                'name': 'Transfer to New Sheet',
                'function': self._execute_different_sheet_transfer
            },
            {
                'id': 'excel_macro',
                'name': 'Execute Excel Macro',
                'function': self._execute_excel_macro
            },
            {
                'id': 'same_sheet_transfer',
                'name': 'Append to Base Sheet',
                'function': self._execute_same_sheet_transfer
            }
        ]

    def update_file_paths(self, base_path, new_data_path):
        """Update the file paths for Excel operations"""
        try:
            # Create backup of base sheet
            if os.path.exists(base_path):
                # Just make a copy of the file to backups folder
                backup_filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(base_path)}"
                backup_path = os.path.join(BACKUP_FOLDER, backup_filename)
                copy2(base_path, backup_path)
                logger.debug(f"Created backup at: {backup_path}")

            self.base_sheet_path = base_path
            self.new_data_sheet_path = new_data_path
            self.output_file_path = os.path.join(OUTPUT_FOLDER, f"processed_{os.path.basename(base_path)}")
            logger.debug(f"Updated file paths - Base: {base_path}, New Data: {new_data_path}, Output: {self.output_file_path}")
        except Exception as e:
            logger.error(f"Error during backup creation: {str(e)}")
            raise

    def execute_pipeline(self, step_order=None):
        """Execute pipeline steps in a separate thread"""
        if not self.base_sheet_path or not self.new_data_sheet_path:
            self.socketio.emit('step_error', {
                'step_id': self.pipeline_steps[0]['id'],
                'error': 'Please upload both Excel files before running the pipeline'
            })
            return

        if step_order:
            # Reorder pipeline steps based on client configuration
            self.pipeline_steps.sort(key=lambda x: step_order.index(x['id']))
        Thread(target=self._run_pipeline).start()

    def _run_pipeline(self):
        """Run all pipeline steps sequentially"""
        for step in self.pipeline_steps:
            try:
                self.socketio.emit('step_start', {'step_id': step['id']})
                step['function']()
                self.socketio.emit('step_complete', {'step_id': step['id']})
            except Exception as e:
                logger.error(f"Error in step {step['id']}: {str(e)}")
                self.socketio.emit('step_error', {
                    'step_id': step['id'],
                    'error': str(e)
                })
                return

        # Copy final result to output folder
        if os.path.exists(self.base_sheet_path):
            try:
                copy2(self.base_sheet_path, self.output_file_path)
                logger.debug(f"Copied final result to: {self.output_file_path}")
                self.socketio.emit('pipeline_complete', {
                    'output_file': os.path.basename(self.output_file_path)
                })
            except Exception as e:
                logger.error(f"Error creating output file: {str(e)}")
                self.socketio.emit('pipeline_error', {
                    'error': f'Failed to generate output file: {str(e)}'
                })
        else:
            logger.error("Output file not found after pipeline execution")
            self.socketio.emit('pipeline_error', {
                'error': 'Failed to generate output file'
            })

    def update_pipeline_config(self, step_order):
        """Update the pipeline configuration based on the new step order"""
        logger.debug(f"Updating pipeline configuration: {step_order}")
        try:
            # Validate step IDs
            if not all(step_id in self.step_functions for step_id in step_order):
                raise ValueError("Invalid step ID in configuration")

            # Reorder pipeline steps
            self.pipeline_steps.sort(key=lambda x: step_order.index(x['id']))
            return True
        except Exception as e:
            logger.error(f"Error updating pipeline configuration: {str(e)}")
            return False

    def _execute_excel_macro(self):
        """Execute Excel macro"""
        from openpyxl import load_workbook
        try:
            logger.debug("Opening workbook")
            workbook = load_workbook(self.base_sheet_path, keep_vba=True)
            # Note: VBA macros cannot be executed in this environment
            # This is a placeholder for macro functionality
            logger.debug("Macro execution skipped - not supported in this environment")
            workbook.save(self.base_sheet_path)
            workbook.close()
            logger.debug("Workbook saved")
        except Exception as e:
            logger.error(f"Macro operation failed: {str(e)}")
            raise

    def _execute_same_sheet_transfer(self):
        """Execute same sheet transfer"""
        try:
            source_workbook = load_workbook(self.new_data_sheet_path)
            source_sheet = source_workbook.active
            destination_workbook = load_workbook(self.base_sheet_path)
            destination_sheet = destination_workbook.active

            next_row = 1
            while destination_sheet.cell(row=next_row, column=1).value is not None:
                next_row += 1

            for row in source_sheet.iter_rows(min_row=2, values_only=True):
                for col_index, value in enumerate(row, start=1):
                    formatted_value = value.strftime('%d/%m/%Y') if isinstance(value, datetime) else value
                    destination_sheet.cell(row=next_row, column=col_index, value=formatted_value)
                next_row += 1

            destination_workbook.save(self.base_sheet_path)
            source_workbook.close()
            destination_workbook.close()
            logger.debug("Same sheet transfer completed")
        except Exception as e:
            logger.error(f"Same sheet transfer failed: {str(e)}")
            raise

    def _execute_different_sheet_transfer(self):
        """Execute different sheet transfer"""
        try:
            source_workbook = load_workbook(self.new_data_sheet_path)
            source_sheet = source_workbook.active
            destination_workbook = load_workbook(self.base_sheet_path)

            new_sheet_name = "Transferred Data"
            destination_sheet = destination_workbook.create_sheet(title=new_sheet_name)

            for row in source_sheet.iter_rows(values_only=True):
                formatted_row = [value.strftime('%d/%m/%Y') if isinstance(value, datetime) else value for value in row]
                destination_sheet.append(formatted_row)

            destination_workbook.save(self.base_sheet_path)
            source_workbook.close()
            destination_workbook.close()
            logger.debug("Different sheet transfer completed")
        except Exception as e:
            logger.error(f"Different sheet transfer failed: {str(e)}")
            raise